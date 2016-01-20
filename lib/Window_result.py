#!/usr/bin/env python
# -*- coding: utf-8 -*-

__author__ = "Jonathan LAMBERT"

from openpyxl import Workbook
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.cell import get_column_letter

import pygtk
pygtk.require('2.0')
import gtk
import string
import os
import time
from Window import Window
import Tools


class Window_result(Window):
	def __init__(self, parent, dictRes):
		## STOCK ITEM
		self.stock_items = (
		("gtk-sendResult", "", 0, 0, None),
		)
		self.stock_aliases = (
		("gtk-sendResult", gtk.STOCK_EXECUTE),
		)
		gtk.stock_add(self.stock_items)
		factory = gtk.IconFactory()
		factory.add_default()
		style = gtk.Style()
		for item, alias in self.stock_aliases:
			icon_set = style.lookup_icon_set(alias)
			factory.add(item, icon_set)
		## FENETRE
		self.parent=parent
		Window.__init__(self, "Administration Oracle : Résultats")
		self.fenetre.connect("delete_event", self.close_window)
		settings = gtk.settings_get_default()
		settings.props.gtk_button_images = True
		self.vbox_main = gtk.VBox(False, 20)
		self.fenetre.add(self.vbox_main)
		self.frm_res=gtk.Frame("Résultats")
		self.scrolled_window = gtk.ScrolledWindow()
		self.scrolled_window.set_border_width(10)
		self.scrolled_window.set_policy(gtk.POLICY_NEVER, gtk.POLICY_NEVER)
		self.scrolled_window.add_with_viewport(self.frm_res)
		self.tab_send = gtk.Table(2, 1, False)
		self.frm_send=gtk.Frame("Envoi des résultats par mail ")
		self.frm_send.add(self.tab_send)
		self.vbox_main.pack_start(self.frm_send, False, False, 0)
		self.vbox_main.pack_start(self.scrolled_window, True, True, 0)
		self.vbox_resultats= gtk.VBox(False, 0)
                self.entry_mail = gtk.Entry()
                self.entry_mail.set_width_chars(50)
		self.entry_mail.set_text("dest1@sigma.fr;dest2@gmail.com;...")
		self.but_mailRes = gtk.Button(stock="gtk-sendResult")
		self.but_mailRes.modify_bg(gtk.STATE_NORMAL, self.but_color)
		self.but_mailRes.connect("clicked", self.sendResult, dictRes)
		self.tab_send.attach(self.entry_mail, 0, 1, 0, 1)
		self.tab_send.attach(self.but_mailRes, 0, 1, 1, 2)
		self.frm_res.add(self.vbox_resultats)
		self.setResults(dictRes)
		self.fenetre.show_all()

		# Min Size
		winSize=self.fenetre.get_size()
		screen = self.fenetre.get_screen()
		if winSize[0]+80 >= screen.get_width()-80:
			newWidth = screen.get_width()-80
		else:
			newWidth = winSize[0]+80
		if winSize[1]+80 >= screen.get_height()-80:
			newHeight = screen.get_height()-80
		else:
			newHeight = winSize[1]+80
		self.fenetre.resize(newWidth, newHeight)
		self.scrolled_window.set_policy(gtk.POLICY_AUTOMATIC, gtk.POLICY_AUTOMATIC)

	def setResults(self, dictRes):
		for dbName, res in dictRes.iteritems():
			cadre=gtk.Frame(dbName)
			numRows = len(res)
			numCols = len(res[0])
			resTab = gtk.Table(numRows, numCols, False)
			cadre.add(resTab)
			for rowInd, line in enumerate(res):
				for colInd, cell in enumerate(line):
					cellContent = None
					cellContent = gtk.Label(str(cell))
					cellContent.set_alignment(xalign=0.0, yalign=0.5)
					resTab.attach(cellContent, colInd, colInd+1, rowInd, rowInd+1)
			self.vbox_resultats.pack_start(cadre, True, True, 10)

	def sendResult(self, widget, dictRes):
		mailForm = self.entry_mail.get_text()
		mailList = mailForm.replace(',',';').strip().split(';')
		if len(mailList) == 0:
			self.error_dialog("\nStatspack-Warning-Message:\nAucune adresse mail n'a été saisie.")
			return
		timestamp = str(time.time())
		self.outFile="/tmp/resultats_"+timestamp+".xlsx"
		self.workBook=Workbook(guess_types=True)
		self.border='medium'
		self.borderColor='FF000000'
		for dbName, res in dictRes.iteritems():
			wsCourante = self.workBook.create_sheet()
			wsCourante.title = 'Resultats - '+dbName
			for indxLine, line in enumerate(res):
				for indxCol, col in enumerate(line):
					indice = string.uppercase[indxLine]+str(indxCol+1)
					wsCourante[string.uppercase[indxLine]+str(indxCol+1)] = col
			# Mise en forme
			for cell in wsCourante.rows[0]:
				cell.font = Font(bold=True)
			column_widths = []
			for row in wsCourante.rows:
				for i, cell in enumerate(row):
					cell.alignment = Alignment(horizontal='center', vertical='center')
					cell.border = Border(left=Side(border_style=self.border, color=self.borderColor), right=Side(border_style=self.border, color=self.borderColor), top=Side(border_style=self.border, color=self.borderColor), bottom=Side(border_style=self.border, color=self.borderColor))
					if len(column_widths) > i:
						if len(str(cell.value)) > column_widths[i]:
							column_widths[i] = len(str(cell.value))
					else:
						column_widths.append(len(str(cell.value)))
			for i, column_width in enumerate(column_widths):
				wsCourante.column_dimensions[get_column_letter(i+1)].width = int((column_width+2)*1.2)
		self.workBook.remove_sheet(self.workBook.worksheets[0])
		self.workBook.save(self.outFile)
		# MAIL
		self.tools=Tools.Tools()
		try:
			self.tools.sendMailWithAttachment("jlambert@sigma.fr",mailList,"Resultats requete oradmin", "Contenu en piece-jointe", [self.outFile])
		except:
			self.warn_dialog("\nMail-Warn-Message:\nProblème d'envoi de mail. Vérifiez les adresses :\n "+", ".join(mailList))
			os.remove(self.outFile)
			return
		os.remove(self.outFile)
		self.notif_dialog("\nQuery-Info-Message:\nRésultats envoyés à "+", ".join(mailList))


	def close_window(self, widget, data=None):
		self.parent.progressBar_workingSQL.set_text("")
		self.parent.progressBar_workingSQL.set_fraction(0.0)
		self.fenetre.destroy()

